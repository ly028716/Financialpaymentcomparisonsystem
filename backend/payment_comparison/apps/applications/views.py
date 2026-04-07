"""
付款申请视图
"""
from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q
from django.utils import timezone
from decimal import Decimal

from payment_comparison.common.response import ApiResponse
from payment_comparison.common.permissions import (
    IsApplicant, IsAccountant, IsFinanceManager, IsAdmin,
    IsOwnerOrAccountantOrAbove, IsAccountantOrFinanceManager,
    CanApproveLargeAmount
)
from .models import PaymentApplication, AuditLog
from .serializers import (
    PaymentApplicationSerializer, PaymentApplicationListSerializer,
    PaymentApplicationDetailSerializer,
    ApproveApplicationSerializer, RejectApplicationSerializer,
    BatchApproveSerializer
)


class ApplicationListAPIView(generics.ListCreateAPIView):
    """
    付款申请列表/创建

    GET: 获取申请列表（根据角色过滤数据）
    POST: 创建付款申请
    """
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return PaymentApplicationSerializer
        return PaymentApplicationListSerializer

    def get_queryset(self):
        user = self.request.user
        queryset = PaymentApplication.objects.all()

        # 根据角色过滤数据
        if user.role == 'applicant':
            # 部门申请人只能看到自己的申请
            queryset = queryset.filter(applicant=user.name, department=user.department)

        # 筛选条件
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        department = self.request.query_params.get('department')
        if department:
            queryset = queryset.filter(department=department)

        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            try:
                from datetime import datetime
                datetime.strptime(start_date, '%Y-%m-%d')
                queryset = queryset.filter(application_date__gte=start_date)
            except ValueError:
                return Response(
                    ApiResponse.error(400, '开始日期格式不正确，应为 YYYY-MM-DD').data,
                    status=status.HTTP_400_BAD_REQUEST
                )
        if end_date:
            try:
                from datetime import datetime
                datetime.strptime(end_date, '%Y-%m-%d')
                queryset = queryset.filter(application_date__lte=end_date)
            except ValueError:
                return Response(
                    ApiResponse.error(400, '结束日期格式不正确，应为 YYYY-MM-DD').data,
                    status=status.HTTP_400_BAD_REQUEST
                )

        keyword = self.request.query_params.get('keyword')
        if keyword:
            # 添加关键词长度限制，防止恶意输入
            if len(keyword) > 100:
                return Response(
                    ApiResponse.error(400, '搜索关键词过长').data,
                    status=status.HTTP_400_BAD_REQUEST
                )
            # 使用 icontains 进行不区分大小写的搜索
            queryset = queryset.filter(
                Q(payee_name__icontains=keyword) |
                Q(application_no__icontains=keyword)
            )

        return queryset.order_by('-created_at')

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return ApiResponse.error(400, '参数错误', serializer.errors)

        application = serializer.save()

        # 记录审核日志
        AuditLog.objects.create(
            application=application,
            action=AuditLog.Action.CREATE,
            operator=request.user.name,
            note='提交付款申请'
        )

        return ApiResponse.created(
            data={
                'id': application.id,
                'application_no': application.application_no,
                'status': application.status,
                'created_at': application.created_at.isoformat()
            },
            message='申请提交成功'
        )

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        serializer = self.get_serializer(page, many=True)
        return ApiResponse.paginated(
            data=serializer.data,
            total=queryset.count(),
            page=request.query_params.get('page', 1),
            page_size=request.query_params.get('page_size', 20)
        )


class MyApplicationListView(generics.ListAPIView):
    """我的申请列表"""
    serializer_class = PaymentApplicationListSerializer
    permission_classes = [IsAuthenticated, IsApplicant]

    def get_queryset(self):
        user = self.request.user
        return PaymentApplication.objects.filter(
            applicant=user.name,
            department=user.department
        ).order_by('-created_at')


class ApplicationDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    """
    付款申请详情

    GET: 获取申请详情
    PUT/PATCH: 更新申请（仅草稿和已拒绝状态可修改）
    DELETE: 撤销申请（仅待审核状态可撤销）
    """
    serializer_class = PaymentApplicationDetailSerializer
    permission_classes = [IsAuthenticated]
    queryset = PaymentApplication.objects.all()

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return ApiResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()

        # 检查是否可编辑
        if not instance.can_edit():
            return ApiResponse.error(403, '当前状态不允许修改')

        serializer = self.get_serializer(instance, data=request.data, partial=True)
        if not serializer.is_valid():
            return ApiResponse.error(400, '参数错误', serializer.errors)

        application = serializer.save()

        # 记录审核日志
        AuditLog.objects.create(
            application=application,
            action=AuditLog.Action.UPDATE,
            operator=request.user.name,
            note='修改付款申请'
        )

        return ApiResponse.success(data=serializer.data, message='修改成功')

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        # 检查是否可撤销
        if not instance.can_delete():
            return ApiResponse.error(403, '当前状态不允许撤销')

        # 记录审核日志
        AuditLog.objects.create(
            application=instance,
            action=AuditLog.Action.CANCEL,
            operator=request.user.name,
            note='撤销付款申请'
        )

        instance.delete()
        return ApiResponse.success(message='申请已撤销')


class PendingApplicationListView(generics.ListAPIView):
    """待审核申请列表"""
    serializer_class = PaymentApplicationListSerializer
    permission_classes = [IsAuthenticated, IsAccountantOrFinanceManager]

    def get_queryset(self):
        queryset = PaymentApplication.objects.filter(
            status=PaymentApplication.Status.PENDING
        )

        # 筛选条件
        department = self.request.query_params.get('department')
        if department:
            queryset = queryset.filter(department=department)

        amount_min = self.request.query_params.get('amount_min')
        amount_max = self.request.query_params.get('amount_max')
        if amount_min:
            try:
                queryset = queryset.filter(amount__gte=Decimal(amount_min))
            except (ValueError, TypeError):
                pass  # 忽略无效的金额参数
        if amount_max:
            try:
                queryset = queryset.filter(amount__lte=Decimal(amount_max))
            except (ValueError, TypeError):
                pass  # 忽略无效的金额参数

        urgent = self.request.query_params.get('urgent')
        if urgent:
            queryset = queryset.filter(urgent=urgent.lower() == 'true')

        # 排序
        sort = self.request.query_params.get('sort', 'created_at')
        if sort == 'amount_desc':
            queryset = queryset.order_by('-amount')
        elif sort == 'amount_asc':
            queryset = queryset.order_by('amount')
        elif sort == 'urgent':
            queryset = queryset.order_by('-urgent', '-created_at')
        else:
            queryset = queryset.order_by('-created_at')

        return queryset


class ApproveApplicationView(APIView):
    """审核通过"""
    permission_classes = [IsAuthenticated, IsAccountantOrFinanceManager]

    def put(self, request, id):
        try:
            application = PaymentApplication.objects.get(id=id)
        except PaymentApplication.DoesNotExist:
            return ApiResponse.error(404, '申请不存在')

        # 检查状态
        if application.status != PaymentApplication.Status.PENDING:
            return ApiResponse.error(400, '申请不在待审核状态')

        # 检查大额审批权限
        LARGE_AMOUNT_THRESHOLD = Decimal('100000')
        if application.amount >= LARGE_AMOUNT_THRESHOLD:
            if request.user.role not in ['finance_manager', 'admin']:
                return ApiResponse.error(403, '大额付款需财务主管审批')

        serializer = ApproveApplicationSerializer(data=request.data)
        if not serializer.is_valid():
            return ApiResponse.error(400, '参数错误', serializer.errors)

        # 更新状态
        application.status = PaymentApplication.Status.APPROVED
        application.save()

        # 记录审核日志
        AuditLog.objects.create(
            application=application,
            action=AuditLog.Action.APPROVE,
            operator=request.user.name,
            note=serializer.validated_data.get('note', '审核通过')
        )

        return ApiResponse.success(
            data={
                'id': application.id,
                'application_no': application.application_no,
                'status': application.status,
                'approved_at': timezone.now().isoformat(),
                'approved_by': request.user.name
            },
            message='审核通过'
        )


class RejectApplicationView(APIView):
    """审核拒绝"""
    permission_classes = [IsAuthenticated, IsAccountantOrFinanceManager]

    def put(self, request, id):
        try:
            application = PaymentApplication.objects.get(id=id)
        except PaymentApplication.DoesNotExist:
            return ApiResponse.error(404, '申请不存在')

        # 检查状态
        if application.status != PaymentApplication.Status.PENDING:
            return ApiResponse.error(400, '申请不在待审核状态')

        serializer = RejectApplicationSerializer(data=request.data)
        if not serializer.is_valid():
            return ApiResponse.error(400, '参数错误', serializer.errors)

        # 更新状态
        application.status = PaymentApplication.Status.REJECTED
        application.remark = serializer.validated_data.get('reason', '')
        application.save()

        # 记录审核日志
        AuditLog.objects.create(
            application=application,
            action=AuditLog.Action.REJECT,
            operator=request.user.name,
            note=serializer.validated_data.get('note', serializer.validated_data['reason'])
        )

        return ApiResponse.success(
            data={
                'id': application.id,
                'application_no': application.application_no,
                'status': application.status,
                'rejected_at': timezone.now().isoformat(),
                'rejected_by': request.user.name,
                'reason': serializer.validated_data['reason']
            },
            message='审核已拒绝'
        )


class BatchApproveView(APIView):
    """批量审核"""
    permission_classes = [IsAuthenticated, IsAccountantOrFinanceManager]

    def post(self, request):
        serializer = BatchApproveSerializer(data=request.data)
        if not serializer.is_valid():
            return ApiResponse.error(400, '参数错误', serializer.errors)

        application_ids = serializer.validated_data['application_ids']
        action = serializer.validated_data['action']
        note = serializer.validated_data.get('note', '')

        results = []
        success_count = 0
        failed_count = 0

        applications = PaymentApplication.objects.filter(
            id__in=application_ids,
            status=PaymentApplication.Status.PENDING
        )

        for app in applications:
            # 检查大额审批权限
            if app.amount >= Decimal('100000') and action == 'approve':
                if request.user.role not in ['finance_manager', 'admin']:
                    results.append({
                        'id': app.id,
                        'application_no': app.application_no,
                        'success': False,
                        'error': '大额付款需财务主管审批'
                    })
                    failed_count += 1
                    continue

            if action == 'approve':
                app.status = PaymentApplication.Status.APPROVED
                audit_action = AuditLog.Action.APPROVE
            else:
                app.status = PaymentApplication.Status.REJECTED
                audit_action = AuditLog.Action.REJECT

            app.save()

            AuditLog.objects.create(
                application=app,
                action=audit_action,
                operator=request.user.name,
                note=note or f'批量{action}'
            )

            results.append({
                'id': app.id,
                'application_no': app.application_no,
                'status': app.status,
                'success': True
            })
            success_count += 1

        return ApiResponse.success(
            data={
                'success_count': success_count,
                'failed_count': failed_count,
                'results': results
            },
            message='批量审核完成'
        )


class ExportApplicationView(APIView):
    """导出付款表"""
    permission_classes = [IsAuthenticated, IsAccountantOrFinanceManager]

    def get(self, request):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side

        # 获取筛选条件
        application_ids = request.query_params.getlist('application_ids')
        status = request.query_params.get('status', 'approved')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        queryset = PaymentApplication.objects.all()

        if application_ids:
            queryset = queryset.filter(id__in=application_ids)
        if status:
            queryset = queryset.filter(status=status)
        if start_date:
            queryset = queryset.filter(application_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(application_date__lte=end_date)

        # 创建Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '付款申请表'

        # 表头
        headers = ['序号', '申请单号', '申请部门', '收款户名', '收款账号', '开户行', '金额', '用途', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 数据
        for row, app in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=app.application_no)
            ws.cell(row=row, column=3, value=app.department)
            ws.cell(row=row, column=4, value=app.payee_name)
            ws.cell(row=row, column=5, value=app.payee_account)
            ws.cell(row=row, column=6, value=app.payee_bank)
            ws.cell(row=row, column=7, value=float(app.amount))
            ws.cell(row=row, column=8, value=app.purpose)
            ws.cell(row=row, column=9, value=app.remark)

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 15

        # 响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payment_applications_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)

        return response